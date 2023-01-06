import json5
import urllib.request
import openpyxl
import logging
from datetime import datetime

#日志配置
_log = logging.getLogger(__name__)
_log.setLevel(logging.INFO)
_ch = logging.StreamHandler()
_ch.setLevel(logging.DEBUG)
_LOG_FORMAT = '%(asctime)s - %(levelname)s - %(funcName)s:%(lineno)d - %(message)s'
_ch.setFormatter(logging.Formatter(_LOG_FORMAT))
_log.addHandler(_ch)

_BANNER_URL="https://github.com/MadeBaruna/paimon-moe/raw/main/src/data/banners.js"
_LOCALE_URL="https://github.com/MadeBaruna/paimon-moe/raw/main/src/locales/items/zh.json"
_ENCODE="utf-8"
_PAIMON_DATE_FORMAT="%Y-%m-%d %H:%M:%S"
_UIGF_DATE_FORMAT=_PAIMON_DATE_FORMAT
_PAIMON_TITLE=["Type","Name","Time","⭐","Pity","#Roll","Group","Banner","Part"]
_GACHA_TYPE_DICT={"100":["Beginners' Wish","新手祈愿"],
                  "200":["Standard", "常驻祈愿"],
                  "301":["Character Event","角色活动祈愿"],
                  "400":["Character Event","角色活动祈愿"],
                  "302":["Weapon Event","武器活动祈愿"]}
_PULL_TYPE_DICT={
    "Weapon":"武器",
    "Character":"角色"
}

def _get_key (dic, value):
    for k, v in dic.items():
         if v == value:
            return k
    _log.warning(f"未找到{value}")
    _log.debug(f"查找{value}堆栈如下：",stack_info=True)
    return None
  
def _get_url(url:str)->str:
    _log.debug(f"获取{url}")
    url=urllib.request.urlopen(url)
    return url.read().decode(_ENCODE)

def get_banners()->dict:
    banner_js_str=_get_url(_BANNER_URL)
    banners_dict=json5.loads(banner_js_str.split('=')[1].split(';')[0])
    for banners in banners_dict.values():
        for wish in banners:
            for wish_key in wish:
                if wish_key != 'start' and wish_key != 'end':
                    continue
                wish[wish_key]=datetime.strptime(wish[wish_key], _PAIMON_DATE_FORMAT)
    return banners_dict

def get_locale()->dict:
    locale_raw = _get_url(_LOCALE_URL)
    return json5.loads(locale_raw)

def get_uigf_from_file(path:str)->dict:
    _log.debug(f"从{path}中获取UIGF")
    with open(path,"r",encoding=_ENCODE) as f:
        uigf_dict=json5.load(f,encoding=_ENCODE)
    return uigf_dict

def _init_workbook()->openpyxl.workbook.Workbook:
    ret=openpyxl.workbook.Workbook()
    ws = ret.active
    titles = [x[0] for x in _GACHA_TYPE_DICT.values()]
    _log.debug(f"工作表标题为：{titles}")
    ws.title = titles[0]
    _log.debug(f"修改第一个工作表为：{titles[0]}")
    ws.append(_PAIMON_TITLE)
    for i in range(1,len(titles)):
        if titles[i] in ret.sheetnames:
            continue
        _log.debug(f"创建工作表：{titles[i]}")
        ret.create_sheet(titles[i])
        ws=ret[titles[i]]
        ws.append(_PAIMON_TITLE)
    return ret

def get_banner_by_date(
        banners_dict:dict,  #banner字典
        time:datetime,        #UIGF时间
        gacha_type:str      #UIGF gacha_type
        )->tuple:           #返回（标题:str，是否切换池子:bool）
    flag = False
    if gacha_type == '302':
        #武器活动祈愿
        _log.debug(f"查找武器活动祈愿：{time}")
        idx = get_banner_by_date.weapon_idx
        banners = banners_dict['weapons']
    else:
        #角色活动祈愿
        _log.debug(f"查找角色活动祈愿（{gacha_type}）：{time}")
        idx = get_banner_by_date.character_idx
        banners = banners_dict['characters']
    while not banners[idx]['start'] < time < banners[idx]['end']:
        flag=True
        if time < banners[idx]['start']:
            _log.warning(f"祈愿{gacha_type}中存在时间{time}的祈愿顺序错误")
            idx-=1
        else :
            idx+=1
    if gacha_type == '302':
        #武器活动祈愿
        get_banner_by_date.weapon_idx=idx
    else:
        #角色活动祈愿
        get_banner_by_date.character_idx = idx
    return banners[idx]['name'],flag
get_banner_by_date.character_idx=0
get_banner_by_date.weapon_idx=0

def _get_valid_col(row):
    ret = 0
    for i in row:
        if i.value is None:
            break
        else:
            ret +=1
    return ret

def fix_pity_row_group_by_sheet(ws:openpyxl.worksheet.worksheet.Worksheet):
    time_col=_PAIMON_TITLE.index("Time")
    star_col=_PAIMON_TITLE.index("⭐")
    pity_col=_PAIMON_TITLE.index("Pity")
    roll_col=_PAIMON_TITLE.index("#Roll")
    group_col=_PAIMON_TITLE.index("Group")
    banner_col=_PAIMON_TITLE.index("Banner")
    max_col=len(_PAIMON_TITLE)
    star5_pity_count = 1
    star4_pity_count = 1
    if ws.max_row > 1:
        ws[2][pity_col].value=1
        ws[2][roll_col].value=1
        ws[2][group_col].value=1
    if ws.max_row < 2:
        return
    for i in range(3,ws.max_row+1):
        #先填group
        if ws[i][time_col].value == ws[i-1][time_col].value:
            #时间相同为一组
            ws[i][group_col].value = ws[i-1][group_col].value
            time_same=True
        else :
            #时间不同为下一组，但是不确定是从1开始还是+1
            time_same = False
        #再比较banner和列数
        if ws[i][banner_col].value == ws[i-1][banner_col].value and\
            _get_valid_col(ws[i])<=max_col :
            #与上一卡池一致
            ws[i][roll_col].value = ws[i-1][roll_col].value+1
            if not time_same:
                ws[i][group_col].value = ws[i-1][group_col].value+1
            if str(ws[i][star_col].value) == "3":
                #抽出3星，Pity列置1，计数+1
                ws[i][pity_col].value = 1
                star5_pity_count+=1
                star4_pity_count+=1            
            elif str(ws[i][star_col].value) == "4":
                #抽出4星
                ws[i][pity_col].value = star4_pity_count
                star5_pity_count+=1
                star4_pity_count=1
            elif str(ws[i][star_col].value) == "5":
                #抽出5星
                ws[i][pity_col].value = star5_pity_count
                star4_pity_count+=1
                star5_pity_count=1
            else:
                #错误
                print(f"无法解析星级{ws[i][star_col].value}")
        else:
            #新卡池
            ws[i][pity_col].value = 1
            ws[i][roll_col].value = 1
            ws[i][group_col].value = 1
            star5_pity_count=1
            star4_pity_count=1

def fix_pity_row_group(wb:openpyxl.workbook.Workbook):
    for sheet_name in wb.sheetnames:
        fix_pity_row_group_by_sheet(wb[sheet_name])

def trim_worksheet(ws:openpyxl.worksheet.worksheet.Worksheet):
    valid_col_num = len(_PAIMON_TITLE)
    if valid_col_num < ws.max_column:
        ws.delete_cols(valid_col_num+1,ws.max_column - valid_col_num)

def trim_workbook(wb:openpyxl.workbook.Workbook):
    for sheet_name in wb.sheetnames:
        trim_worksheet(wb[sheet_name])


def add_extra_data(wb:openpyxl.workbook.Workbook,banners_dict:dict):
    #添加Banner List Sheet
    banner_sheet_name = "Banner List"
    banner_sheet_titles=["Name","Start","End"]
    wb.create_sheet(banner_sheet_name)
    ws=wb[banner_sheet_name]
    ws.append(banner_sheet_titles)
    #新手祈愿
    ws.append((
        banners_dict['beginners'][0]['name'],
        banners_dict['beginners'][0]['start'].strftime(_PAIMON_DATE_FORMAT),
        banners_dict['beginners'][0]['end'].strftime(_PAIMON_DATE_FORMAT)))
    #常驻祈愿
    ws.append((
        banners_dict['standard'][0]['name'],
        banners_dict['standard'][0]['start'].strftime(_PAIMON_DATE_FORMAT),
        banners_dict['standard'][0]['end'].strftime(_PAIMON_DATE_FORMAT)))
    #角色活动祈愿
    for event in banners_dict['characters']:
        ws.append((
            event['name'],
            event['start'].strftime(_PAIMON_DATE_FORMAT),
            event['end'].strftime(_PAIMON_DATE_FORMAT),
        ))
    #武器活动祈愿
    for event in banners_dict['weapons']:
        ws.append((
            event['name'],
            event['start'].strftime(_PAIMON_DATE_FORMAT),
            event['end'].strftime(_PAIMON_DATE_FORMAT),
        ))
    
    #添加Information Sheet
    info_sheet_name = "Information"
    wb.create_sheet(info_sheet_name)
    ws=wb[info_sheet_name]
    ws['A1'] = "Paimon.moe Wish History Export"
    ws['A2'] = 'Version'
    ws['B2'] = '3'
    ws['A3'] = 'Export Date'
    ws['B3'] = datetime.now().strftime(_PAIMON_DATE_FORMAT)
    ws.merge_cells('A1:B1')

def main():
    _log.info("开始读取祈愿信息")
    banners_dict=get_banners()
    _log.info("祈愿信息读取完毕")
    _log.debug(f"祈愿信息字典：\n{banners_dict}")
    _log.info("开始读取本地化信息")
    locale_dict = get_locale()
    _log.info("本地化信息读取完毕")
    _log.debug(f"本地化信息字典：\n{banners_dict}")
    _log.info("开始读取UIGF信息")
    uigf_dict=get_uigf_from_file("uigf.json")
    _log.info("UIGF信息读取完毕")
    _log.debug(f"UIGF信息字典：\n{uigf_dict}")
    workbook=_init_workbook()
    _log.info(f"开始将UIGF信息转换为Excel")
    for gacha in uigf_dict['list']:
        worksheet=workbook[_GACHA_TYPE_DICT[gacha['gacha_type']][0]]
        row=[]
        #Type
        row.append(_get_key(_PULL_TYPE_DICT, gacha['item_type']))
        #Name
        name=_get_key(locale_dict, gacha['name'])
        if name == None:
            print(f"{gacha['name']}未找到对应翻译")
            row.append("")
        else:
            row.append(name)
        #Time
        time=datetime.strptime(gacha['time'], _UIGF_DATE_FORMAT)
        row.append(time.strftime(_PAIMON_DATE_FORMAT))
        #⭐
        row.append(gacha['rank_type'])
        #Pity 暂时不在这里添加
        row.append("")
        ##Roll 暂时不在这里添加
        row.append("")
        #Group 暂时不在这里添加
        row.append("")
        #Banner
        diff_flag = False
        if gacha['gacha_type'] == "100":
            row.append(banners_dict['beginners'][0]['name'])
        elif gacha['gacha_type'] == "200":
            row.append(banners_dict['standard'][0]['name'])
        else:
            banner, diff_flag = get_banner_by_date(banners_dict, time, gacha['gacha_type'])
            row.append(banner)
        #Part
        if gacha['gacha_type'] == "400":
            row.append("Wish 2")
        else :
            row.append("")
        if diff_flag == True:
            row.append("1")
        worksheet.append(row)
    _log.info(f"信息转换完毕，开始填充Pity、Row、Group列")
    fix_pity_row_group(workbook)
    _log.info(f"填充完毕，开始清理Excel文件")
    trim_workbook(workbook)
    _log.info(f"清理Excel文件完毕，开始添加额外信息")
    add_extra_data(workbook, banners_dict)
    _log.info(f"添加额外信息完毕，保存文件")
    workbook.save("test_output.xlsx")
    _log.info(f"保存完毕，执行结束")

if __name__=="__main__":
    main()
